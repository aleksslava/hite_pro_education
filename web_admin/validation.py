from __future__ import annotations

import html
import re
from html.parser import HTMLParser
from io import BytesIO
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook


NAME_MASK = "[Имя]"
MAX_XLSX_SIZE = 10 * 1024 * 1024
TELEGRAM_TEXT_LIMIT = 4096
TELEGRAM_CAPTION_LIMIT = 1024
MAX_TEXT_LIMIT = 4000
ALLOWED_ACTIONS = {
    "main_menu": "Главное меню",
    "stat": "Статистика обучения",
    **{f"lesson_{number}": f"Урок {number}" for number in range(1, 8)},
    "exam": "Экзамен",
}
ALLOWED_TAGS = {
    "b", "strong", "i", "em", "u", "ins", "s", "strike", "del",
    "span", "tg-spoiler", "a", "code", "pre", "blockquote",
}
ALLOWED_ENTITIES = {"lt", "gt", "amp", "quot"}
LANGUAGE_CLASS_RE = re.compile(r"language-[a-zA-Z0-9_+.#-]+\Z")


class UploadValidationError(ValueError):
    pass


class _TelegramHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=False)
        self.stack: list[str] = []
        self.text_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.casefold()
        if tag not in ALLOWED_TAGS:
            raise UploadValidationError(f"Тег <{tag}> не поддерживается Telegram.")
        if "code" in self.stack:
            raise UploadValidationError("Внутри <code> нельзя использовать другие теги.")
        if "pre" in self.stack and tag != "code":
            raise UploadValidationError("Внутри <pre> допускается только тег <code>.")
        if tag == "a" and "a" in self.stack:
            raise UploadValidationError("Ссылки <a> нельзя вкладывать друг в друга.")
        self._validate_attrs(tag, attrs)
        self.stack.append(tag)

    def handle_startendtag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        raise UploadValidationError(f"Самозакрывающийся тег <{tag}/> не поддерживается.")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.casefold()
        if not self.stack or self.stack[-1] != tag:
            expected = self.stack[-1] if self.stack else "нет"
            raise UploadValidationError(
                f"Некорректное закрытие </{tag}>; ожидается </{expected}>."
            )
        self.stack.pop()

    def handle_data(self, data: str) -> None:
        if any(character in data for character in "<>&"):
            raise UploadValidationError(
                "Символы <, > и & вне HTML-тегов нужно записывать как &lt;, &gt; и &amp;."
            )
        self.text_parts.append(data)

    def handle_entityref(self, name: str) -> None:
        if name not in ALLOWED_ENTITIES:
            raise UploadValidationError(f"HTML-сущность &{name}; не поддерживается Telegram.")
        self.text_parts.append(html.unescape(f"&{name};"))

    def handle_charref(self, name: str) -> None:
        try:
            value = int(name[1:], 16) if name.lower().startswith("x") else int(name)
            self.text_parts.append(chr(value))
        except (ValueError, OverflowError) as error:
            raise UploadValidationError(f"Некорректная HTML-сущность: &#{name};") from error

    def handle_comment(self, data: str) -> None:
        raise UploadValidationError("HTML-комментарии в сообщении не поддерживаются.")

    def close(self) -> None:
        super().close()
        if self.stack:
            raise UploadValidationError(f"Не закрыт тег <{self.stack[-1]}>.")

    def _validate_attrs(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_map = {name.casefold(): value for name, value in attrs}
        if len(attrs_map) != len(attrs):
            raise UploadValidationError(f"В теге <{tag}> повторяются атрибуты.")
        if tag == "a":
            if set(attrs_map) != {"href"} or not attrs_map["href"]:
                raise UploadValidationError("Тег <a> должен содержать только атрибут href.")
            parsed = urlparse(attrs_map["href"])
            if parsed.scheme not in {"http", "https", "tg"}:
                raise UploadValidationError("Ссылка должна использовать http, https или tg://.")
            if parsed.scheme in {"http", "https"} and not parsed.netloc:
                raise UploadValidationError("Укажите полный адрес ссылки.")
            if parsed.scheme == "tg" and not attrs_map["href"].startswith("tg://user?id="):
                raise UploadValidationError("Разрешены только ссылки tg://user?id=…")
            return
        if tag == "span":
            if attrs_map != {"class": "tg-spoiler"}:
                raise UploadValidationError('Для <span> разрешён только class="tg-spoiler".')
            return
        if tag == "blockquote":
            if attrs_map not in ({}, {"expandable": None}, {"expandable": ""}):
                raise UploadValidationError("Для <blockquote> разрешён только атрибут expandable.")
            return
        if tag == "code" and attrs_map:
            if (
                set(attrs_map) != {"class"}
                or "pre" not in self.stack
                or not attrs_map["class"]
                or not LANGUAGE_CLASS_RE.fullmatch(attrs_map["class"])
            ):
                raise UploadValidationError(
                    'Атрибут class="language-…" разрешён только для <code> внутри <pre>.'
                )
            return
        if attrs_map:
            raise UploadValidationError(f"Атрибуты тега <{tag}> не поддерживаются.")

    @property
    def visible_text(self) -> str:
        return "".join(self.text_parts)


def validate_telegram_html(message: str, *, limit: int | None = None) -> str:
    if not message.strip():
        raise UploadValidationError("Введите текст сообщения.")
    parser = _TelegramHTMLParser()
    try:
        parser.feed(message)
        parser.close()
    except UploadValidationError:
        raise
    except Exception as error:
        raise UploadValidationError("Не удалось разобрать HTML-разметку.") from error
    visible = parser.visible_text
    if not visible.strip():
        raise UploadValidationError("После удаления HTML-тегов сообщение пусто.")
    if limit is not None and len(visible) > limit:
        raise UploadValidationError(
            f"Видимый текст должен быть не длиннее {limit} символов."
        )
    return visible


def render_message(message: str, name: str) -> str:
    return message.replace(NAME_MASK, html.escape(name, quote=False))


class _MaxHTMLRenderer(HTMLParser):
    """Render already validated Telegram HTML into the MAX HTML subset."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=False)
        self.output: list[str] = []
        self.stack: list[str | None] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.casefold()
        attrs_map = {name.casefold(): value for name, value in attrs}
        emitted: str | None
        if tag in {"tg-spoiler", "span"}:
            emitted = None
        elif tag == "strike":
            emitted = "s"
        elif tag == "a":
            href = attrs_map.get("href", "") or ""
            if urlparse(href).scheme in {"http", "https"}:
                emitted = "a"
                self.output.append(f'<a href="{html.escape(href, quote=True)}">')
                self.stack.append(emitted)
                return
            emitted = None
        else:
            emitted = tag
        if emitted is not None:
            self.output.append(f"<{emitted}>")
        self.stack.append(emitted)

    def handle_endtag(self, tag: str) -> None:
        emitted = self.stack.pop()
        if emitted is not None:
            self.output.append(f"</{emitted}>")

    def handle_data(self, data: str) -> None:
        self.output.append(data)

    def handle_entityref(self, name: str) -> None:
        self.output.append(f"&{name};")

    def handle_charref(self, name: str) -> None:
        self.output.append(f"&#{name};")


class _MaxHTMLValidator(HTMLParser):
    TAGS = {"b", "strong", "i", "em", "u", "ins", "s", "del", "a", "code", "pre", "blockquote"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=False)
        self.stack: list[str] = []
        self.text_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.casefold()
        if tag not in self.TAGS:
            raise UploadValidationError(f"Тег <{tag}> не поддерживается MAX.")
        attrs_map = {name.casefold(): value for name, value in attrs}
        if tag == "a":
            href = attrs_map.get("href")
            parsed = urlparse(href or "")
            if set(attrs_map) != {"href"} or parsed.scheme not in {"http", "https"} or not parsed.netloc:
                raise UploadValidationError("MAX поддерживает только полные HTTP(S)-ссылки.")
        elif attrs_map:
            raise UploadValidationError(f"Атрибуты тега <{tag}> не поддерживаются MAX.")
        self.stack.append(tag)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.casefold()
        if not self.stack or self.stack[-1] != tag:
            raise UploadValidationError("Некорректное закрытие HTML-тега для MAX.")
        self.stack.pop()

    def handle_data(self, data: str) -> None:
        self.text_parts.append(data)

    def handle_entityref(self, name: str) -> None:
        if name not in ALLOWED_ENTITIES:
            raise UploadValidationError(f"HTML-сущность &{name}; не поддерживается MAX.")
        self.text_parts.append(html.unescape(f"&{name};"))

    def handle_charref(self, name: str) -> None:
        try:
            value = int(name[1:], 16) if name.lower().startswith("x") else int(name)
            self.text_parts.append(chr(value))
        except (ValueError, OverflowError) as error:
            raise UploadValidationError(f"Некорректная HTML-сущность: &#{name};") from error

    def close(self) -> None:
        super().close()
        if self.stack:
            raise UploadValidationError(f"Не закрыт тег <{self.stack[-1]}> для MAX.")


def adapt_telegram_html_for_max(message: str, *, limit: int = MAX_TEXT_LIMIT) -> str:
    validate_telegram_html(message)
    renderer = _MaxHTMLRenderer()
    renderer.feed(message)
    renderer.close()
    rendered = "".join(renderer.output)
    validator = _MaxHTMLValidator()
    validator.feed(rendered)
    validator.close()
    visible = "".join(validator.text_parts)
    if not visible.strip():
        raise UploadValidationError("После адаптации для MAX сообщение пусто.")
    if len(visible) > limit:
        raise UploadValidationError(f"Текст для MAX должен быть не длиннее {limit} символов.")
    return rendered


def normalize_recipient_id(value: Any) -> int:
    if isinstance(value, bool):
        raise ValueError
    if isinstance(value, int):
        recipient_id = value
    elif isinstance(value, float) and value.is_integer():
        recipient_id = int(value)
    else:
        text = str(value or "").strip()
        if not text.isdigit():
            raise ValueError
        recipient_id = int(text)
    if recipient_id <= 0:
        raise ValueError
    return recipient_id


def parse_recipients(
    file_content: bytes,
    message: str,
    *,
    message_limit: int,
    targets: set[str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    targets = targets or {"telegram"}
    if not targets or not targets.issubset({"telegram", "max"}):
        raise UploadValidationError("Выберите Telegram и/или MAX.")
    if not file_content:
        raise UploadValidationError("Excel-файл пуст.")
    if len(file_content) > MAX_XLSX_SIZE:
        raise UploadValidationError("Excel-файл должен быть не больше 10 МБ.")
    try:
        workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception as error:
        raise UploadValidationError("Не удалось открыть Excel-файл.") from error
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise UploadValidationError("В Excel-файле нет заголовков.")
        columns = {
            str(value).strip().casefold(): index
            for index, value in enumerate(headers)
            if value is not None
        }
        missing = [name for name in ("telegram_id", "max_id", "имя") if name not in columns]
        if missing:
            raise UploadValidationError(f"Не найдены обязательные колонки: {', '.join(missing)}.")

        recipients: list[dict[str, Any]] = []
        seen = {"telegram": set(), "max": set()}
        platform_stats = {
            platform: {"ready": 0, "skipped": 0, "duplicates": 0, "invalid": 0}
            for platform in targets
        }
        stats: dict[str, Any] = {
            "ready": 0,
            "skipped": 0,
            "duplicates": 0,
            "invalid": 0,
            "platforms": platform_stats,
        }
        needs_name = NAME_MASK in message
        for row_number, row in enumerate(rows, start=2):
            if all(value in (None, "") for value in row):
                continue
            raw_tg = row[columns["telegram_id"]] if columns["telegram_id"] < len(row) else None
            raw_max = row[columns["max_id"]] if columns["max_id"] < len(row) else None
            raw_name = row[columns["имя"]] if columns["имя"] < len(row) else None
            name = str(raw_name or "").strip()
            try:
                telegram_id = normalize_recipient_id(raw_tg)
            except ValueError:
                telegram_id = None
            try:
                max_id = normalize_recipient_id(raw_max)
            except ValueError:
                max_id = None
            deliveries: dict[str, dict[str, Any]] = {}
            for platform in targets:
                target_id = telegram_id if platform == "telegram" else max_id
                raw_target = raw_tg if platform == "telegram" else raw_max
                status, error, issue = "pending", None, None
                if target_id is None:
                    status, error, issue = "skipped", f"Некорректный {'telegram_id' if platform == 'telegram' else 'max_id'}", "invalid"
                elif target_id in seen[platform]:
                    status, error, issue = "skipped", f"Повторный {'telegram_id' if platform == 'telegram' else 'max_id'}", "duplicates"
                elif needs_name and not name:
                    status, error, issue = "skipped", "Не указано имя для шаблона [Имя]", "invalid"
                else:
                    try:
                        personalized = render_message(message, name)
                        if platform == "telegram":
                            validate_telegram_html(personalized, limit=message_limit)
                        else:
                            adapt_telegram_html_for_max(personalized, limit=MAX_TEXT_LIMIT)
                    except UploadValidationError as validation_error:
                        status, error, issue = "skipped", str(validation_error), "invalid"

                if status == "pending":
                    seen[platform].add(target_id)
                    platform_stats[platform]["ready"] += 1
                    stats["ready"] += 1
                else:
                    platform_stats[platform]["skipped"] += 1
                    stats["skipped"] += 1
                    if issue:
                        platform_stats[platform][issue] += 1
                        stats[issue] += 1
                deliveries[platform] = {
                    "target_id": target_id,
                    "raw_target_id": str(raw_target or "").strip(),
                    "status": status,
                    "error": error,
                }
            primary = deliveries["telegram"] if "telegram" in deliveries else deliveries["max"]
            recipients.append({
                "row_number": row_number,
                "telegram_id": telegram_id,
                "raw_telegram_id": str(raw_tg or "").strip(),
                "max_id": max_id,
                "raw_max_id": str(raw_max or "").strip(),
                "name": name,
                "status": primary["status"],
                "error": primary["error"],
                "deliveries": deliveries,
            })
    finally:
        workbook.close()
    if not recipients:
        raise UploadValidationError("В Excel-файле нет получателей.")
    if not stats["ready"]:
        raise UploadValidationError("В Excel-файле нет корректных получателей.")
    return recipients, stats


def validate_buttons(buttons: list[dict[str, str]]) -> list[dict[str, str]]:
    if len(buttons) > 8:
        raise UploadValidationError("В сообщении может быть не больше восьми кнопок.")
    result: list[dict[str, str]] = []
    for button in buttons:
        text = button.get("text", "").strip()
        action = button.get("action_key", "").strip()
        if not text:
            continue
        if not action:
            raise UploadValidationError("Для каждой кнопки укажите текст и действие.")
        if len(text) > 64:
            raise UploadValidationError("Текст кнопки должен быть не длиннее 64 символов.")
        if action not in ALLOWED_ACTIONS:
            raise UploadValidationError("Выбрано неизвестное действие кнопки.")
        result.append({"text": text, "action_key": action})
    return result
